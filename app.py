import os
import json
import uuid
import hashlib
import re
import sqlite3
import threading
import time
import tkinter as tk
from tkinter import filedialog
from datetime import datetime
from typing import Any, TypedDict
from flask import Flask, render_template, request, jsonify, Response
from flask_cors import CORS
import torch
import easyocr
import whisper
import ollama
import psutil
from langgraph.graph import StateGraph, END
from langchain_ollama import ChatOllama
from openpyxl import load_workbook

# Get the absolute path of the directory where this script is located
project_root = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=project_root
)
CORS(app)

# --- Configuration & Model Loading ---
UPLOAD_FOLDER = 'uploads'
HISTORY_FILE = 'history.json'
MODEL_NAMES_FILE = 'model_names.json'
CUSTOMER_DB_FILE = os.path.join(project_root, 'customer_data.db')
CONFIG_CUSTOMER_DATA_ROOT = 'customer_data_root'
CONFIG_CUSTOMER_EXCEL_PATH = 'customer_excel_path'
CONFIG_MONITOR_INTERVAL_SECONDS = 'monitor_interval_seconds'
SUPPORTED_IMAGE_EXTENSIONS = {'.png', '.jpg', '.jpeg', '.webp', '.bmp'}
SUPPORTED_AUDIO_EXTENSIONS = {'.mp3', '.wav', '.m4a', '.aac', '.ogg', '.flac'}
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

monitor_lock = threading.Lock()
monitor_stop_event = threading.Event()
monitor_thread: threading.Thread | None = None
monitor_runtime = {
    'thread_running': False,
    'monitor_active': False,
    'scan_in_progress': False,
    'scan_phase': 'idle',
    'current_folder': '',
    'current_file': '',
    'last_scan_at': None,
    'last_error': '',
    'last_result_message': '',
    'last_scan_new_records': 0,
    'last_scan_processed_files': 0,
    'last_scan_duplicates_skipped': 0,
}

print("INFO: Loading skill guidelines...")
with open(os.path.join(project_root, 'skill.md'), 'r', encoding='utf-8') as f:
    SKILL_GUIDELINES = f.read()

# Determine device
DEVICE = "cuda" if torch.cuda.is_available() else "cpu"
print(f"INFO: Using device: {DEVICE}")

# --- Model Management ---
# Use a dictionary to cache loaded models
loaded_models = {}

def get_whisper_model(model_name: str):
    """Loads a Whisper model into memory, caching it for future use."""
    if model_name in loaded_models:
        print(f"INFO: Using cached Whisper model '{model_name}'.")
        return loaded_models[model_name]

    # Simple cache eviction: unload all other models to make space
    if loaded_models:
        print("INFO: Unloading existing models to free up memory.")
        loaded_models.clear()
        if DEVICE == 'cuda':
            torch.cuda.empty_cache()

    print(f"INFO: Loading Whisper model '{model_name}'...")
    model = whisper.load_model(model_name, device=DEVICE)
    print(f"INFO: Whisper model '{model_name}' loaded.")
    loaded_models[model_name] = model
    return model

# Pre-load the base model on startup
get_whisper_model("base")

print("INFO: Loading EasyOCR reader...")
ocr_reader = easyocr.Reader(['en'], gpu=(DEVICE == 'cuda'))

print("INFO: Flask app and models loaded successfully.")

def load_model_names():
    """Loads the available model names and descriptions from JSON file."""
    try:
        with open(MODEL_NAMES_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        # Return empty lists on error
        return {"agent_models": [], "whisper_models": []}

# --- History Management ---
def load_history():
    """Loads the analysis history from the JSON file."""
    if not os.path.exists(HISTORY_FILE):
        return []
    try:
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except (json.JSONDecodeError, IOError):
        return []

def save_history(data):
    """Saves the updated analysis history to the JSON file."""
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4)


def normalize_header(value: Any) -> str:
    return re.sub(r'[^a-z0-9]+', '_', str(value or '').strip().lower()).strip('_')


def normalize_phone_number(value: Any) -> str:
    return re.sub(r'\D+', '', str(value or ''))


def is_supported_media_file(file_name: str) -> bool:
    extension = os.path.splitext(file_name)[1].lower()
    return extension in SUPPORTED_IMAGE_EXTENSIONS or extension in SUPPORTED_AUDIO_EXTENSIONS


def parse_summary_fields(summary_text: str) -> dict[str, str]:
    fields = {
        'Customer Name': 'N/A',
        'Issue Category': 'N/A',
        'Key Information': 'N/A',
        'Payment Date & Time': 'N/A',
        'Payment Amount': 'N/A',
        'Payer Details': 'N/A',
        'Payee Details': 'N/A',
        'UPI Transaction ID': 'N/A',
        'Recommended Next Step': 'N/A',
    }
    if not summary_text:
        return fields

    pattern = re.compile(r'\*\*(.+?):\*\*\s*(.*?)(?=\n\*\*|$)', re.S)
    for raw_key, raw_value in pattern.findall(summary_text):
        key = raw_key.strip()
        if key in fields:
            fields[key] = raw_value.strip()
    return fields


def load_monitor_config() -> dict[str, Any]:
    config = load_model_names()
    return {
        'customer_data_root': str(config.get(CONFIG_CUSTOMER_DATA_ROOT, '') or '').strip(),
        'customer_excel_path': str(config.get(CONFIG_CUSTOMER_EXCEL_PATH, '') or '').strip(),
        'monitor_interval_seconds': int(config.get(CONFIG_MONITOR_INTERVAL_SECONDS, 15) or 15),
        'default_agent_model': config.get('default_agent_model', 'llama3'),
        'default_whisper': config.get('default_whisper', 'base'),
    }


def get_db_connection() -> sqlite3.Connection:
    connection = sqlite3.connect(CUSTOMER_DB_FILE, timeout=30)
    connection.row_factory = sqlite3.Row
    connection.execute('PRAGMA journal_mode=WAL')
    connection.execute('PRAGMA busy_timeout = 30000')
    return connection


def ensure_column_exists(connection: sqlite3.Connection, table_name: str, column_name: str, column_sql: str) -> None:
    existing_columns = {
        row['name'] for row in connection.execute(f'PRAGMA table_info({table_name})').fetchall()
    }
    if column_name not in existing_columns:
        connection.execute(f'ALTER TABLE {table_name} ADD COLUMN {column_sql}')


def init_customer_database() -> None:
    with get_db_connection() as connection:
        connection.executescript(
            """
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_name TEXT NOT NULL,
                phone_number TEXT,
                normalized_phone TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                UNIQUE(customer_name, normalized_phone)
            );

            CREATE TABLE IF NOT EXISTS accounts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                account_name TEXT,
                account_number TEXT,
                folder_name TEXT NOT NULL UNIQUE,
                source_folder_path TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id)
            );

            CREATE TABLE IF NOT EXISTS artifacts (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                customer_id INTEGER,
                account_id INTEGER,
                folder_name TEXT,
                source_file_path TEXT NOT NULL UNIQUE,
                file_name TEXT,
                file_type TEXT,
                issue_category TEXT,
                specialist_path TEXT,
                summary TEXT,
                transcription TEXT,
                ocr_text TEXT,
                proposed_action TEXT,
                action_status TEXT,
                action_result TEXT,
                payment_amount TEXT,
                payment_datetime TEXT,
                payer_details TEXT,
                payee_details TEXT,
                upi_transaction_id TEXT,
                processed_at TEXT NOT NULL,
                FOREIGN KEY(customer_id) REFERENCES customers(id),
                FOREIGN KEY(account_id) REFERENCES accounts(id)
            );

            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                artifact_id INTEGER,
                customer_id INTEGER,
                account_id INTEGER,
                payment_amount TEXT,
                payment_datetime TEXT,
                payer_details TEXT,
                payee_details TEXT,
                upi_transaction_id TEXT UNIQUE,
                matched_on TEXT,
                created_at TEXT NOT NULL,
                FOREIGN KEY(artifact_id) REFERENCES artifacts(id),
                FOREIGN KEY(customer_id) REFERENCES customers(id),
                FOREIGN KEY(account_id) REFERENCES accounts(id)
            );
            """
        )
        ensure_column_exists(connection, 'artifacts', 'content_hash', 'content_hash TEXT')
        connection.execute('CREATE INDEX IF NOT EXISTS idx_artifacts_content_hash ON artifacts(content_hash)')


def compute_file_sha256(file_path: str) -> str:
    digest = hashlib.sha256()
    with open(file_path, 'rb') as source_file:
        while True:
            chunk = source_file.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def load_customer_mapping_from_excel(excel_path: str) -> dict[str, dict[str, str]]:
    if not excel_path or not os.path.exists(excel_path):
        return {}

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    header_row = next(rows, None)
    if not header_row:
        workbook.close()
        return {}

    headers = [normalize_header(value) for value in header_row]

    def find_index(options: list[str]) -> int | None:
        for option in options:
            if option in headers:
                return headers.index(option)
        return None

    folder_index = find_index(['folder_name', 'folder', 'customer_folder', 'shared_folder', 'data_folder'])
    customer_name_index = find_index(['customer_name', 'customer', 'name'])
    phone_index = find_index(['phone_number', 'phone', 'mobile', 'mobile_number'])
    account_number_index = find_index(['account_number', 'account_no', 'account', 'account_id'])
    account_name_index = find_index(['account_name', 'account_label', 'branch', 'location', 'site_name'])

    mappings: dict[str, dict[str, str]] = {}
    for row in rows:
        if folder_index is None or folder_index >= len(row):
            continue
        folder_value = str(row[folder_index] or '').strip()
        if not folder_value:
            continue
        folder_name = os.path.basename(folder_value.rstrip('\\/'))
        mappings[folder_name.lower()] = {
            'folder_name': folder_name,
            'customer_name': str(row[customer_name_index] or '').strip() if customer_name_index is not None and customer_name_index < len(row) else '',
            'phone_number': str(row[phone_index] or '').strip() if phone_index is not None and phone_index < len(row) else '',
            'account_number': str(row[account_number_index] or '').strip() if account_number_index is not None and account_number_index < len(row) else '',
            'account_name': str(row[account_name_index] or '').strip() if account_name_index is not None and account_name_index < len(row) else '',
        }

    workbook.close()
    return mappings


def open_native_dialog(dialog_type: str) -> str:
    """Open a native file or folder picker on the local machine running the Flask app."""
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    try:
        if dialog_type == 'folder':
            selected_path = filedialog.askdirectory(title='Select Customer Data Folder')
        elif dialog_type == 'excel':
            selected_path = filedialog.askopenfilename(
                title='Select Customer Mapping Excel',
                filetypes=[('Excel files', '*.xlsx *.xlsm *.xltx *.xltm'), ('All files', '*.*')],
            )
        else:
            selected_path = ''
    finally:
        root.destroy()
    return selected_path or ''


def upsert_customer_account(connection: sqlite3.Connection, folder_name: str, folder_path: str, mapping: dict[str, str] | None) -> tuple[int | None, int | None]:
    if not mapping:
        return None, None

    customer_name = mapping.get('customer_name') or folder_name
    phone_number = mapping.get('phone_number', '')
    normalized_phone = normalize_phone_number(phone_number)
    now = datetime.now().isoformat()

    existing_customer = connection.execute(
        'SELECT id FROM customers WHERE customer_name = ? AND normalized_phone = ?',
        (customer_name, normalized_phone),
    ).fetchone()

    if existing_customer:
        customer_id = existing_customer['id']
        connection.execute(
            'UPDATE customers SET phone_number = ?, updated_at = ? WHERE id = ?',
            (phone_number, now, customer_id),
        )
    else:
        cursor = connection.execute(
            'INSERT INTO customers (customer_name, phone_number, normalized_phone, created_at, updated_at) VALUES (?, ?, ?, ?, ?)',
            (customer_name, phone_number, normalized_phone, now, now),
        )
        customer_id = cursor.lastrowid

    existing_account = connection.execute(
        'SELECT id FROM accounts WHERE folder_name = ?',
        (folder_name,),
    ).fetchone()

    account_name = mapping.get('account_name') or folder_name
    account_number = mapping.get('account_number', '')
    if existing_account:
        account_id = existing_account['id']
        connection.execute(
            'UPDATE accounts SET customer_id = ?, account_name = ?, account_number = ?, source_folder_path = ?, updated_at = ? WHERE id = ?',
            (customer_id, account_name, account_number, folder_path, now, account_id),
        )
    else:
        cursor = connection.execute(
            'INSERT INTO accounts (customer_id, account_name, account_number, folder_name, source_folder_path, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)',
            (customer_id, account_name, account_number, folder_name, folder_path, now, now),
        )
        account_id = cursor.lastrowid

    return customer_id, account_id


def build_initial_state(
    *,
    audio_path: str | None,
    screenshot_path: str | None,
    audio_filename: str,
    screenshot_filename: str,
    model_name: str,
    agent_model: str,
    require_human_review: bool,
    approval_granted: bool,
    persist_history: bool,
) -> dict[str, Any]:
    return {
        'audio_path': audio_path,
        'screenshot_path': screenshot_path,
        'audio_filename': audio_filename,
        'screenshot_filename': screenshot_filename,
        'model_name': model_name,
        'agent_model': agent_model,
        'require_human_review': require_human_review,
        'approval_granted': approval_granted,
        'persist_history': persist_history,
        'ocr_result': [],
        'transcription_text': '',
        'issue_category': 'Other Issues',
        'specialist_path': 'General support specialist',
        'summary_text': '',
        'needs_human_review': False,
        'review_reason': '',
        'proposed_action': '',
        'action_payload': {},
        'action_status': 'not_started',
        'action_result': '',
        'error': '',
    }


def save_monitored_artifact(
    connection: sqlite3.Connection,
    *,
    folder_name: str,
    file_path: str,
    content_hash: str,
    file_type: str,
    final_state: dict[str, Any],
    customer_id: int | None,
    account_id: int | None,
) -> None:
    summary_fields = parse_summary_fields(final_state.get('summary_text', ''))
    now = datetime.now().isoformat()
    cursor = connection.execute(
        '''
        INSERT OR IGNORE INTO artifacts (
            customer_id, account_id, folder_name, source_file_path, file_name, file_type,
            issue_category, specialist_path, summary, transcription, ocr_text,
            proposed_action, action_status, action_result,
            payment_amount, payment_datetime, payer_details, payee_details,
            upi_transaction_id, content_hash, processed_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''',
        (
            customer_id,
            account_id,
            folder_name,
            file_path,
            os.path.basename(file_path),
            file_type,
            final_state.get('issue_category', 'Other Issues'),
            final_state.get('specialist_path', 'General support specialist'),
            final_state.get('summary_text', ''),
            final_state.get('transcription_text', ''),
            '\n'.join(final_state.get('ocr_result', [])),
            final_state.get('proposed_action', ''),
            final_state.get('action_status', 'not_started'),
            final_state.get('action_result', ''),
            summary_fields.get('Payment Amount', 'N/A'),
            summary_fields.get('Payment Date & Time', 'N/A'),
            summary_fields.get('Payer Details', 'N/A'),
            summary_fields.get('Payee Details', 'N/A'),
            summary_fields.get('UPI Transaction ID', 'N/A'),
            content_hash,
            now,
        ),
    )

    artifact_id = cursor.lastrowid
    transaction_id = summary_fields.get('UPI Transaction ID', 'N/A')
    payment_amount = summary_fields.get('Payment Amount', 'N/A')
    payment_datetime = summary_fields.get('Payment Date & Time', 'N/A')
    if artifact_id and (transaction_id != 'N/A' or payment_amount != 'N/A' or payment_datetime != 'N/A'):
        connection.execute(
            '''
            INSERT OR IGNORE INTO payments (
                artifact_id, customer_id, account_id, payment_amount, payment_datetime,
                payer_details, payee_details, upi_transaction_id, matched_on, created_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''',
            (
                artifact_id,
                customer_id,
                account_id,
                payment_amount,
                payment_datetime,
                summary_fields.get('Payer Details', 'N/A'),
                summary_fields.get('Payee Details', 'N/A'),
                None if transaction_id == 'N/A' else transaction_id,
                folder_name,
                now,
            ),
        )


def collect_monitor_stats() -> dict[str, int]:
    if not os.path.exists(CUSTOMER_DB_FILE):
        return {'customers': 0, 'accounts': 0, 'artifacts': 0, 'payments': 0}

    with get_db_connection() as connection:
        return {
            'customers': connection.execute('SELECT COUNT(*) FROM customers').fetchone()[0],
            'accounts': connection.execute('SELECT COUNT(*) FROM accounts').fetchone()[0],
            'artifacts': connection.execute('SELECT COUNT(*) FROM artifacts').fetchone()[0],
            'payments': connection.execute('SELECT COUNT(*) FROM payments').fetchone()[0],
        }


def is_artifact_already_processed(file_path: str, content_hash: str | None = None) -> bool:
    with get_db_connection() as connection:
        if content_hash:
            row = connection.execute(
                'SELECT 1 FROM artifacts WHERE content_hash = ? OR source_file_path = ?',
                (content_hash, file_path),
            ).fetchone()
        else:
            row = connection.execute(
                'SELECT 1 FROM artifacts WHERE source_file_path = ?',
                (file_path,),
            ).fetchone()
    return row is not None


def update_monitor_runtime(**kwargs: Any) -> None:
    with monitor_lock:
        monitor_runtime.update(kwargs)


def scan_customer_data_root() -> dict[str, int]:
    init_customer_database()
    config = load_monitor_config()
    root_path = config['customer_data_root']
    excel_path = config['customer_excel_path']
    if not root_path:
        return {'processed_files': 0, 'new_records': 0}
    if not os.path.isdir(root_path):
        raise FileNotFoundError(f'Customer data folder not found: {root_path}')

    mappings = load_customer_mapping_from_excel(excel_path)
    processed_files = 0
    new_records = 0
    duplicates_skipped = 0
    update_monitor_runtime(
        scan_in_progress=True,
        scan_phase='loading_mappings',
        current_folder='',
        current_file='',
        last_error='',
        last_result_message='Reading customer mapping and scanning folders...',
        last_scan_duplicates_skipped=0,
    )

    for entry in os.scandir(root_path):
        if not entry.is_dir():
            continue
        folder_name = entry.name
        update_monitor_runtime(scan_phase='scanning_folder', current_folder=folder_name, current_file='')
        folder_mapping = mappings.get(folder_name.lower())

        with get_db_connection() as connection:
            customer_id, account_id = upsert_customer_account(connection, folder_name, entry.path, folder_mapping)
            connection.commit()

        for dirpath, _, filenames in os.walk(entry.path):
            for file_name in filenames:
                if not is_supported_media_file(file_name):
                    continue
                processed_files += 1
                source_file_path = os.path.join(dirpath, file_name)
                update_monitor_runtime(scan_phase='checking_file', current_folder=folder_name, current_file=source_file_path)
                content_hash = compute_file_sha256(source_file_path)
                if is_artifact_already_processed(source_file_path, content_hash):
                    duplicates_skipped += 1
                    update_monitor_runtime(last_result_message=f'Skipped duplicate file: {os.path.basename(source_file_path)}')
                    continue

                extension = os.path.splitext(file_name)[1].lower()
                is_audio = extension in SUPPORTED_AUDIO_EXTENSIONS
                update_monitor_runtime(scan_phase='processing_file', current_folder=folder_name, current_file=source_file_path)
                final_state = ANALYSIS_GRAPH.invoke(
                    build_initial_state(
                        audio_path=source_file_path if is_audio else None,
                        screenshot_path=source_file_path if not is_audio else None,
                        audio_filename=file_name if is_audio else 'N/A',
                        screenshot_filename=file_name if not is_audio else 'N/A',
                        model_name=config['default_whisper'],
                        agent_model=config['default_agent_model'],
                        require_human_review=False,
                        approval_granted=False,
                        persist_history=False,
                    )
                )

                with get_db_connection() as connection:
                    save_monitored_artifact(
                        connection,
                        folder_name=folder_name,
                        file_path=source_file_path,
                        content_hash=content_hash,
                        file_type='audio' if is_audio else 'image',
                        final_state=final_state,
                        customer_id=customer_id,
                        account_id=account_id,
                    )
                    connection.commit()
                new_records += 1

    update_monitor_runtime(
        scan_in_progress=False,
        scan_phase='idle',
        current_folder='',
        current_file='',
        last_result_message=f'Scan complete. {new_records} new records, {duplicates_skipped} duplicates skipped from {processed_files} supported files.',
        last_scan_duplicates_skipped=duplicates_skipped,
    )

    return {'processed_files': processed_files, 'new_records': new_records, 'duplicates_skipped': duplicates_skipped}


def customer_monitor_loop() -> None:
    while not monitor_stop_event.is_set():
        config = load_monitor_config()
        interval = max(5, int(config['monitor_interval_seconds']))
        try:
            update_monitor_runtime(thread_running=True, monitor_active=bool(config['customer_data_root']))

            scan_result = scan_customer_data_root() if config['customer_data_root'] else {'processed_files': 0, 'new_records': 0}
            update_monitor_runtime(
                last_scan_at=datetime.now().isoformat(),
                last_error='',
                last_scan_processed_files=scan_result['processed_files'],
                last_scan_new_records=scan_result['new_records'],
                last_scan_duplicates_skipped=scan_result.get('duplicates_skipped', 0),
            )
        except Exception as e:
            update_monitor_runtime(
                scan_in_progress=False,
                scan_phase='error',
                current_folder='',
                current_file='',
                last_scan_at=datetime.now().isoformat(),
                last_error=str(e),
                last_result_message='Monitor scan failed.',
            )
        monitor_stop_event.wait(interval)


def start_customer_monitor() -> None:
    global monitor_thread
    if monitor_thread and monitor_thread.is_alive():
        return
    monitor_stop_event.clear()
    monitor_thread = threading.Thread(target=customer_monitor_loop, daemon=True, name='customer-data-monitor')
    monitor_thread.start()


def _dummy_activate_account(payload: dict[str, Any]) -> dict[str, Any]:
    """Dummy billing action placeholder. Replace with real activation endpoint call later."""
    return {
        "status": "executed",
        "message": "Dummy activation executed for billing/deactivation issue.",
        "payload": payload
    }


def _dummy_schedule_technician(payload: dict[str, Any]) -> dict[str, Any]:
    """Dummy connectivity action placeholder. Replace with field dispatch endpoint."""
    return {
        "status": "executed",
        "message": "Dummy technician scheduling action executed for connectivity issue.",
        "payload": payload
    }


def _dummy_process_plan_change(payload: dict[str, Any]) -> dict[str, Any]:
    """Dummy plan change action placeholder. Replace with CRM/plan endpoint."""
    return {
        "status": "executed",
        "message": "Dummy plan-change workflow executed.",
        "payload": payload
    }


def _dummy_create_service_ticket(payload: dict[str, Any]) -> dict[str, Any]:
    """Dummy fallback action placeholder. Replace with service desk endpoint."""
    return {
        "status": "executed",
        "message": "Dummy service-ticket creation executed.",
        "payload": payload
    }


ACTION_HANDLERS: dict[str, Any] = {
    "activate_account": _dummy_activate_account,
    "schedule_technician": _dummy_schedule_technician,
    "process_plan_change": _dummy_process_plan_change,
    "create_service_ticket": _dummy_create_service_ticket,
}


def execute_planned_action(action_name: str, payload: dict[str, Any]) -> dict[str, Any]:
    """Executes an action using dummy handlers. Swap handlers with real endpoints later."""
    handler = ACTION_HANDLERS.get(action_name, _dummy_create_service_ticket)
    try:
        return handler(payload)
    except Exception as e:
        return {
            "status": "failed",
            "message": f"Action {action_name} failed: {str(e)}",
            "payload": payload,
        }


class AgentState(TypedDict):
    """State shared across the LangGraph workflow."""
    audio_path: str | None
    screenshot_path: str | None
    audio_filename: str
    screenshot_filename: str
    model_name: str
    agent_model: str
    require_human_review: bool
    approval_granted: bool
    persist_history: bool
    ocr_result: list[str]
    transcription_text: str
    issue_category: str
    specialist_path: str
    summary_text: str
    needs_human_review: bool
    review_reason: str
    proposed_action: str
    action_payload: dict[str, Any]
    action_status: str
    action_result: str
    error: str


def build_analysis_graph():
    """Builds and compiles the analysis agent workflow using LangGraph."""

    def preprocess_node(state: AgentState) -> dict[str, Any]:
        return {
            "ocr_result": [],
            "transcription_text": "",
            "issue_category": "Other Issues",
            "specialist_path": "General support specialist",
            "summary_text": "",
            "needs_human_review": False,
            "review_reason": "",
            "proposed_action": "",
            "action_payload": {},
            "action_status": "not_started",
            "action_result": "",
            "error": ""
        }

    def ocr_node(state: AgentState) -> dict[str, Any]:
        screenshot_path = state.get("screenshot_path")
        if not screenshot_path:
            return {"ocr_result": []}
        ocr_result = ocr_reader.readtext(screenshot_path, detail=0, paragraph=True)
        return {"ocr_result": ocr_result}

    def transcribe_node(state: AgentState) -> dict[str, Any]:
        audio_path = state.get("audio_path")
        if not audio_path:
            return {"transcription_text": ""}
        whisper_model = get_whisper_model(state.get("model_name", "base"))
        audio_transcription = whisper_model.transcribe(audio_path, fp16=(DEVICE == 'cuda'))
        return {"transcription_text": audio_transcription.get("text", "")}

    def classify_issue_node(state: AgentState) -> dict[str, Any]:
        text_blob = f"{state.get('transcription_text', '')} {' '.join(state.get('ocr_result', []))}".lower()
        billing_tokens = [
            "bill", "billing", "payment", "upi", "invoice", "due", "recharge", "deactivated", "account"
        ]
        connectivity_tokens = [
            "internet", "connection", "wifi", "router", "slow", "disconnect", "los", "wan"
        ]
        plan_tokens = ["plan", "upgrade", "downgrade", "speed", "offer"]

        if any(token in text_blob for token in billing_tokens):
            category = "Billing Issue / Account Deactivated"
            specialist_path = "Billing specialist"
        elif any(token in text_blob for token in connectivity_tokens):
            category = "Connectivity Issue"
            specialist_path = "Connectivity and field-support specialist"
        elif any(token in text_blob for token in plan_tokens):
            category = "Plan Change / Upgrade Request"
            specialist_path = "General support specialist"
        else:
            category = "Other Issues"
            specialist_path = "General support specialist"

        return {"issue_category": category, "specialist_path": specialist_path}

    def build_prompt(state: AgentState, specialist_focus: str) -> str:
        transcription_text = state.get("transcription_text", "")
        ocr_result = state.get("ocr_result", [])
        issue_category = state.get("issue_category", "Other Issues")

        analysis_request = f"""
**Customer Call Analysis Request**

**Audio Transcription:**
{transcription_text if transcription_text else "N/A"}

**Text Extracted from Screenshot:**
{' '.join(ocr_result) if ocr_result else "N/A"}

**Routed Issue Category:**
{issue_category}
"""

        return f"""
You are an AI assistant for Interjet, a high-speed internet provider.
You are currently acting as: {specialist_focus}
Your task is to analyze the provided customer interaction and generate a summary based on the company's standard operating procedures.
If only a screenshot is provided, focus on analyzing the details from the image.
If only audio is provided, focus on the transcription.
If both are provided, use both as context.

Follow these guidelines strictly:
{SKILL_GUIDELINES}

---
Here is the interaction data to analyze:
{analysis_request}
---

Please provide the English summary now. Structure your response using the following markdown format, ensuring each field is on a new line:
**Customer Name:** [Customer's name from audio or screenshot, or N/A]
**Issue Category:** [The categorized issue]
**Key Information:** [Other key details from the call or text, NOT related to the payment]
**Payment Date & Time:** [Date and Time from screenshot, e.g., "July 13, 2026, 11:03 AM", or N/A]
**Payment Amount:** [Amount from screenshot as a number only, e.g., 599, or N/A]
**Payer Details:** [Payer name or UPI ID from screenshot, or N/A]
**Payee Details:** [Payee name or UPI ID from screenshot, or N/A]
**UPI Transaction ID:** [Transaction ID from screenshot, or N/A]
**Recommended Next Step:** [The recommended next step based on the SOPs]
"""

    def run_summary(state: AgentState, specialist_focus: str) -> dict[str, Any]:
        agent_model = state.get("agent_model", "llama3")
        llm = ChatOllama(model=agent_model, temperature=0)
        llm_response = llm.invoke(build_prompt(state, specialist_focus))
        summary_text = getattr(llm_response, "content", "") or ""
        return {"summary_text": summary_text}

    def summarize_billing_node(state: AgentState) -> dict[str, Any]:
        return run_summary(state, "Billing specialist")

    def summarize_connectivity_node(state: AgentState) -> dict[str, Any]:
        return run_summary(state, "Connectivity and field-support specialist")

    def summarize_general_node(state: AgentState) -> dict[str, Any]:
        return run_summary(state, "General support specialist")

    def review_gate_node(state: AgentState) -> dict[str, Any]:
        require_review = state.get("require_human_review", False)
        issue_category = state.get("issue_category", "Other Issues")
        ocr_text = " ".join(state.get("ocr_result", [])).lower()
        summary_text = state.get("summary_text", "").lower()

        if not require_review:
            return {"needs_human_review": False, "review_reason": ""}

        # Ask for explicit human confirmation on billing cases with weak payment identifiers.
        weak_payment_evidence = (
            issue_category == "Billing Issue / Account Deactivated"
            and not any(token in ocr_text for token in ["transaction", "upi", "reference", "ref no", "txn"])
            and "upi transaction id" in summary_text
            and "n/a" in summary_text
        )

        if weak_payment_evidence:
            return {
                "needs_human_review": True,
                "review_reason": "Billing flow detected but transaction identifiers were not confidently extracted."
            }
        return {"needs_human_review": False, "review_reason": ""}

    def human_review_note_node(state: AgentState) -> dict[str, Any]:
        reason = state.get("review_reason", "Manual verification requested.")
        summary = state.get("summary_text", "")
        review_note = f"\n\n**Human Review Required:** {reason}"
        return {"summary_text": f"{summary}{review_note}"}

    def propose_action_node(state: AgentState) -> dict[str, Any]:
        issue_category = state.get("issue_category", "Other Issues")
        needs_human_review = state.get("needs_human_review", False)

        action_map = {
            "Billing Issue / Account Deactivated": "activate_account",
            "Connectivity Issue": "schedule_technician",
            "Plan Change / Upgrade Request": "process_plan_change",
            "Other Issues": "create_service_ticket"
        }
        proposed_action = action_map.get(issue_category, "create_service_ticket")
        action_payload = {
            "issue_category": issue_category,
            "specialist_path": state.get("specialist_path", "General support specialist"),
            "review_flag": needs_human_review,
            "review_reason": state.get("review_reason", ""),
            "audio_file": state.get("audio_filename", "N/A"),
            "screenshot_file": state.get("screenshot_filename", "N/A"),
        }
        return {
            "proposed_action": proposed_action,
            "action_payload": action_payload,
            "action_status": "planned",
            "action_result": f"Planned action: {proposed_action}"
        }

    def approval_gate_node(state: AgentState) -> dict[str, Any]:
        if state.get("approval_granted", False):
            return {"action_status": "approved", "action_result": "Approval received. Action execution allowed."}

        reason = "Approval not provided from UI/API request."
        if state.get("needs_human_review", False):
            reason = state.get("review_reason", reason)
        return {
            "action_status": "awaiting_approval",
            "action_result": f"Execution skipped. {reason}"
        }

    def execute_action_node(state: AgentState) -> dict[str, Any]:
        action_name = state.get("proposed_action", "create_service_ticket")
        result = execute_planned_action(action_name, state.get("action_payload", {}))
        return {
            "action_status": result.get("status", "executed"),
            "action_result": result.get("message", f"Action {action_name} executed.")
        }

    def save_history_node(state: AgentState) -> dict[str, Any]:
        if not state.get("persist_history", True):
            return {}
        try:
            new_entry = {
                "id": f"job-{int(datetime.now().timestamp())}",
                "timestamp": datetime.now().isoformat(),
                "audio_file": state.get("audio_filename", "N/A") or "N/A",
                "screenshot_file": state.get("screenshot_filename", "N/A") or "N/A",
                "issue_category": state.get("issue_category", "Other Issues"),
                "specialist_path": state.get("specialist_path", "General support specialist"),
                "needs_human_review": state.get("needs_human_review", False),
                "review_reason": state.get("review_reason", ""),
                "proposed_action": state.get("proposed_action", ""),
                "action_payload": state.get("action_payload", {}),
                "action_status": state.get("action_status", "not_started"),
                "action_result": state.get("action_result", ""),
                "summary": state.get("summary_text", ""),
                "transcription": state.get("transcription_text", ""),
                "ocr": state.get("ocr_result", [])
            }
            history = load_history()
            history.append(new_entry)
            save_history(history)
        except Exception as e:
            print(f"Warning: Failed to save history entry. Reason: {e}")
        return {}

    workflow = StateGraph(AgentState)
    workflow.add_node("preprocess", preprocess_node)
    workflow.add_node("ocr", ocr_node)
    workflow.add_node("transcribe", transcribe_node)
    workflow.add_node("classify_issue", classify_issue_node)
    workflow.add_node("summarize_billing", summarize_billing_node)
    workflow.add_node("summarize_connectivity", summarize_connectivity_node)
    workflow.add_node("summarize_general", summarize_general_node)
    workflow.add_node("review_gate", review_gate_node)
    workflow.add_node("human_review_note", human_review_note_node)
    workflow.add_node("propose_action", propose_action_node)
    workflow.add_node("approval_gate", approval_gate_node)
    workflow.add_node("execute_action", execute_action_node)
    workflow.add_node("save_history", save_history_node)

    workflow.set_entry_point("preprocess")
    workflow.add_edge("preprocess", "ocr")
    workflow.add_edge("ocr", "transcribe")
    workflow.add_edge("transcribe", "classify_issue")

    def route_by_issue(state: AgentState) -> str:
        category = state.get("issue_category", "Other Issues")
        if category == "Billing Issue / Account Deactivated":
            return "billing"
        if category == "Connectivity Issue":
            return "connectivity"
        return "general"

    workflow.add_conditional_edges(
        "classify_issue",
        route_by_issue,
        {
            "billing": "summarize_billing",
            "connectivity": "summarize_connectivity",
            "general": "summarize_general"
        }
    )

    workflow.add_edge("summarize_billing", "review_gate")
    workflow.add_edge("summarize_connectivity", "review_gate")
    workflow.add_edge("summarize_general", "review_gate")

    def route_review(state: AgentState) -> str:
        return "human_review" if state.get("needs_human_review", False) else "save"

    workflow.add_conditional_edges(
        "review_gate",
        route_review,
        {
            "human_review": "human_review_note",
            "save": "propose_action"
        }
    )

    workflow.add_edge("human_review_note", "propose_action")

    workflow.add_edge("propose_action", "approval_gate")

    def route_action_execution(state: AgentState) -> str:
        return "execute" if state.get("action_status") == "approved" else "save"

    workflow.add_conditional_edges(
        "approval_gate",
        route_action_execution,
        {
            "execute": "execute_action",
            "save": "save_history"
        }
    )

    workflow.add_edge("execute_action", "save_history")
    workflow.add_edge("save_history", END)

    return workflow.compile()


ANALYSIS_GRAPH = build_analysis_graph()

@app.route('/api/history', methods=['GET'])
def get_history():
    """Endpoint to retrieve the analysis history."""
    history = load_history()
    return jsonify(sorted(history, key=lambda x: x['timestamp'], reverse=True))


@app.route('/api/monitor/status', methods=['GET'])
def monitor_status():
    """Returns customer-data monitor status and SQLite counters."""
    with monitor_lock:
        runtime_snapshot = dict(monitor_runtime)
    config = load_monitor_config()
    return jsonify({
        'config': config,
        'database_path': CUSTOMER_DB_FILE,
        'stats': collect_monitor_stats(),
        'runtime': runtime_snapshot,
    })


@app.route('/api/monitor/scan-now', methods=['POST'])
def monitor_scan_now():
    """Runs a one-off monitor scan immediately."""
    try:
        result = scan_customer_data_root()
        with monitor_lock:
            monitor_runtime['last_scan_at'] = datetime.now().isoformat()
            monitor_runtime['last_error'] = ''
            monitor_runtime['last_scan_processed_files'] = result['processed_files']
            monitor_runtime['last_scan_new_records'] = result['new_records']
        return jsonify({
            'message': 'Monitor scan completed.',
            'result': result,
            'stats': collect_monitor_stats(),
        })
    except Exception as e:
        with monitor_lock:
            monitor_runtime['last_scan_at'] = datetime.now().isoformat()
            monitor_runtime['last_error'] = str(e)
        return jsonify({'error': str(e)}), 500


@app.route('/api/monitor/artifacts', methods=['GET'])
def monitor_artifacts():
    """Returns monitored artifact rows from SQLite for dashboard display."""
    if not os.path.exists(CUSTOMER_DB_FILE):
        return jsonify([])

    query = '''
        SELECT
            artifacts.id,
            artifacts.folder_name,
            artifacts.file_name,
            artifacts.file_type,
            artifacts.issue_category,
            artifacts.specialist_path,
            artifacts.proposed_action,
            artifacts.action_status,
            artifacts.action_result,
            artifacts.payment_amount,
            artifacts.payment_datetime,
            artifacts.upi_transaction_id,
            artifacts.processed_at,
            customers.customer_name,
            customers.phone_number,
            accounts.account_name,
            accounts.account_number
        FROM artifacts
        LEFT JOIN customers ON customers.id = artifacts.customer_id
        LEFT JOIN accounts ON accounts.id = artifacts.account_id
        ORDER BY datetime(artifacts.processed_at) DESC
        LIMIT 200
    '''
    with get_db_connection() as connection:
        rows = connection.execute(query).fetchall()
    return jsonify([dict(row) for row in rows])


@app.route('/api/monitor/payments', methods=['GET'])
def monitor_payments():
    """Returns payment matches from SQLite for dashboard display."""
    if not os.path.exists(CUSTOMER_DB_FILE):
        return jsonify([])

    query = '''
        SELECT
            payments.id,
            payments.payment_amount,
            payments.payment_datetime,
            payments.payer_details,
            payments.payee_details,
            payments.upi_transaction_id,
            payments.matched_on,
            payments.created_at,
            customers.customer_name,
            customers.phone_number,
            accounts.account_name,
            accounts.account_number,
            artifacts.file_name,
            artifacts.issue_category
        FROM payments
        LEFT JOIN customers ON customers.id = payments.customer_id
        LEFT JOIN accounts ON accounts.id = payments.account_id
        LEFT JOIN artifacts ON artifacts.id = payments.artifact_id
        ORDER BY datetime(payments.created_at) DESC
        LIMIT 200
    '''
    with get_db_connection() as connection:
        rows = connection.execute(query).fetchall()
    return jsonify([dict(row) for row in rows])


@app.route('/api/monitor/artifacts/<int:artifact_id>/approve-action', methods=['POST'])
def approve_monitor_artifact_action(artifact_id):
    """Approves and executes the planned action for a monitored artifact."""
    with get_db_connection() as connection:
        artifact = connection.execute(
            '''
            SELECT artifacts.*, customers.customer_name, customers.phone_number,
                   accounts.account_name, accounts.account_number
            FROM artifacts
            LEFT JOIN customers ON customers.id = artifacts.customer_id
            LEFT JOIN accounts ON accounts.id = artifacts.account_id
            WHERE artifacts.id = ?
            ''',
            (artifact_id,),
        ).fetchone()

        if not artifact:
            return jsonify({'error': 'Monitored artifact not found.'}), 404

        current_status = artifact['action_status'] or 'not_started'
        if current_status == 'executed':
            return jsonify({'message': 'Action already executed.', 'item': dict(artifact)}), 200
        if current_status == 'rejected':
            return jsonify({'error': 'Action was rejected and cannot be approved until reprocessed.'}), 409

        proposed_action = artifact['proposed_action'] or 'create_service_ticket'
        payload = {
            'artifact_id': artifact['id'],
            'customer_name': artifact['customer_name'] or '',
            'phone_number': artifact['phone_number'] or '',
            'account_name': artifact['account_name'] or '',
            'account_number': artifact['account_number'] or '',
            'folder_name': artifact['folder_name'] or '',
            'file_name': artifact['file_name'] or '',
            'issue_category': artifact['issue_category'] or 'Other Issues',
            'specialist_path': artifact['specialist_path'] or 'General support specialist',
            'payment_amount': artifact['payment_amount'] or '',
            'payment_datetime': artifact['payment_datetime'] or '',
            'upi_transaction_id': artifact['upi_transaction_id'] or '',
        }
        result = execute_planned_action(proposed_action, payload)
        connection.execute(
            'UPDATE artifacts SET action_status = ?, action_result = ? WHERE id = ?',
            (result.get('status', 'executed'), result.get('message', f'Action {proposed_action} executed.'), artifact_id),
        )
        connection.commit()

        updated = connection.execute('SELECT * FROM artifacts WHERE id = ?', (artifact_id,)).fetchone()
    return jsonify({'message': 'Monitored action approved and processed.', 'item': dict(updated)}), 200


@app.route('/api/monitor/artifacts/<int:artifact_id>/reject-action', methods=['POST'])
def reject_monitor_artifact_action(artifact_id):
    """Rejects the planned action for a monitored artifact."""
    request_data = request.get_json(silent=True) or {}
    reason = (request_data.get('reason') or 'Rejected in monitored dashboard flow.').strip()

    with get_db_connection() as connection:
        artifact = connection.execute('SELECT * FROM artifacts WHERE id = ?', (artifact_id,)).fetchone()
        if not artifact:
            return jsonify({'error': 'Monitored artifact not found.'}), 404
        if (artifact['action_status'] or '') == 'executed':
            return jsonify({'error': 'Action already executed and cannot be rejected.'}), 409

        connection.execute(
            'UPDATE artifacts SET action_status = ?, action_result = ? WHERE id = ?',
            ('rejected', f'Rejected: {reason}', artifact_id),
        )
        connection.commit()
        updated = connection.execute('SELECT * FROM artifacts WHERE id = ?', (artifact_id,)).fetchone()
    return jsonify({'message': 'Monitored action rejected.', 'item': dict(updated)}), 200


@app.route('/api/browse/folder', methods=['POST'])
def browse_folder():
    """Open a native folder picker and return the selected path."""
    try:
        selected_path = open_native_dialog('folder')
        return jsonify({'path': selected_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/browse/excel', methods=['POST'])
def browse_excel():
    """Open a native file picker for the customer mapping Excel file."""
    try:
        selected_path = open_native_dialog('excel')
        return jsonify({'path': selected_path})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/history/<item_id>', methods=['DELETE'])
def delete_history_item(item_id):
    """Deletes a specific history item by its ID."""
    history = load_history()
    item_found = any(item['id'] == item_id for item in history)
    if not item_found:
        return jsonify({"error": "Item not found"}), 404

    updated_history = [item for item in history if item['id'] != item_id]
    save_history(updated_history)
    return jsonify({"message": "Item deleted successfully"}), 200


@app.route('/api/history/<item_id>/approve-action', methods=['POST'])
def approve_history_action(item_id):
    """Approves and executes a previously planned action for a history item."""
    history = load_history()
    target_item = next((item for item in history if item.get('id') == item_id), None)
    if not target_item:
        return jsonify({"error": "Item not found"}), 404

    current_status = target_item.get('action_status', 'not_started')
    if current_status == 'executed':
        return jsonify({"message": "Action already executed.", "item": target_item}), 200
    if current_status == 'rejected':
        return jsonify({"error": "Action was rejected. Clear rejection before approving."}), 409

    action_map = {
        "Billing Issue / Account Deactivated": "activate_account",
        "Connectivity Issue": "schedule_technician",
        "Plan Change / Upgrade Request": "process_plan_change",
        "Other Issues": "create_service_ticket"
    }
    proposed_action = target_item.get('proposed_action') or action_map.get(
        target_item.get('issue_category', 'Other Issues'),
        'create_service_ticket'
    )
    payload = target_item.get('action_payload') or {
        "issue_category": target_item.get("issue_category", "Other Issues"),
        "specialist_path": target_item.get("specialist_path", "General support specialist"),
        "review_flag": target_item.get("needs_human_review", False),
        "review_reason": target_item.get("review_reason", ""),
        "audio_file": target_item.get("audio_file", "N/A"),
        "screenshot_file": target_item.get("screenshot_file", "N/A"),
    }

    result = execute_planned_action(proposed_action, payload)
    target_item['proposed_action'] = proposed_action
    target_item['action_payload'] = payload
    target_item['action_status'] = result.get('status', 'executed')
    target_item['action_result'] = result.get('message', f"Action {proposed_action} executed.")
    target_item['approval_state'] = 'approved'
    target_item['approved_at'] = datetime.now().isoformat()

    save_history(history)
    return jsonify({"message": "Action approved and processed.", "item": target_item}), 200


@app.route('/api/history/<item_id>/reject-action', methods=['POST'])
def reject_history_action(item_id):
    """Rejects a previously planned action for a history item."""
    history = load_history()
    target_item = next((item for item in history if item.get('id') == item_id), None)
    if not target_item:
        return jsonify({"error": "Item not found"}), 404

    if target_item.get('action_status') == 'executed':
        return jsonify({"error": "Action already executed and cannot be rejected."}), 409

    request_data = request.get_json(silent=True) or {}
    reason = (request_data.get('reason') or "Rejected in dashboard approval flow.").strip()

    target_item['action_status'] = 'rejected'
    target_item['action_result'] = f"Rejected: {reason}"
    target_item['approval_state'] = 'rejected'
    target_item['rejected_at'] = datetime.now().isoformat()

    save_history(history)
    return jsonify({"message": "Action rejected.", "item": target_item}), 200

@app.route('/')
def index():
    """Renders the main user interface."""
    return render_template('index.html')

@app.route('/api/config', methods=['GET', 'POST'])
def app_config():
    """Endpoint to get or set application configuration."""
    if request.method == 'POST':
        # Load current model data
        model_config = load_model_names()
        # Update only the default keys from the request
        new_defaults = request.json
        model_config['default_agent_model'] = new_defaults.get('default_agent_model', model_config.get('default_agent_model'))
        model_config['default_whisper'] = new_defaults.get('default_whisper', model_config.get('default_whisper'))
        model_config[CONFIG_CUSTOMER_DATA_ROOT] = new_defaults.get(CONFIG_CUSTOMER_DATA_ROOT, model_config.get(CONFIG_CUSTOMER_DATA_ROOT, ''))
        model_config[CONFIG_CUSTOMER_EXCEL_PATH] = new_defaults.get(CONFIG_CUSTOMER_EXCEL_PATH, model_config.get(CONFIG_CUSTOMER_EXCEL_PATH, ''))
        model_config[CONFIG_MONITOR_INTERVAL_SECONDS] = int(new_defaults.get(CONFIG_MONITOR_INTERVAL_SECONDS, model_config.get(CONFIG_MONITOR_INTERVAL_SECONDS, 15)) or 15)
        # Save the entire file back
        with open(MODEL_NAMES_FILE, 'w', encoding='utf-8') as f:
            json.dump(model_config, f, indent=4)
        return jsonify({"message": "Configuration saved successfully."}), 200
    else:
        # Return only the default model keys
        config = load_model_names()
        return jsonify({
            "default_agent_model": config.get("default_agent_model"),
            "default_whisper": config.get("default_whisper"),
            CONFIG_CUSTOMER_DATA_ROOT: config.get(CONFIG_CUSTOMER_DATA_ROOT, ''),
            CONFIG_CUSTOMER_EXCEL_PATH: config.get(CONFIG_CUSTOMER_EXCEL_PATH, ''),
            CONFIG_MONITOR_INTERVAL_SECONDS: config.get(CONFIG_MONITOR_INTERVAL_SECONDS, 15),
        })

@app.route('/api/available-models', methods=['GET'])
def available_models():
    """Endpoint to get the list of available models for UI dropdowns."""
    return jsonify(load_model_names())

@app.route('/api/models-status', methods=['GET'])
def models_status():
    """Checks and returns the status of available AI models."""
    agent_status, whisper_status = {}, {}
    model_names = load_model_names()
    agent_models_to_check = [m['id'] for m in model_names.get('agent_models', [])]
    whisper_models_to_check = [m['id'] for m in model_names.get('whisper_models', [])]

    # Helper function to ensure everything has a tag for exact matching
    def normalize_tag(name):
        if name and ':' not in name:
            return f"{name}:latest"
        return name

    # --- Check Ollama Models ---
    try:
        response = ollama.list()
        ollama_models_on_disk = set()
        
        # Handle dictionary response (Older ollama-python versions)
        if isinstance(response, dict):
            if 'error' in response:
                raise Exception(response['error'])
            ollama_models_on_disk = {model.get('name', model.get('model')) for model in response.get('models', [])}
        
        # Handle object response (Newer ollama-python versions >= 0.2.0)
        else:
            ollama_models_on_disk = {model.model for model in getattr(response, 'models', [])}

        # Normalize all disk models to 'name:tag' format
        normalized_disk_models = {normalize_tag(m) for m in ollama_models_on_disk if m}

        if not normalized_disk_models:
            # If no models are on disk, mark all as not_downloaded
            for model_name in agent_models_to_check:
                agent_status[model_name] = {"status": "not_downloaded"}
        else:
            for model_name in agent_models_to_check:
                # Direct exact check against normalized strings
                is_downloaded = normalize_tag(model_name) in normalized_disk_models
                agent_status[model_name] = {"status": "downloaded" if is_downloaded else "not_downloaded"}
            
    except Exception as e:
        # Broad exception catches httpx.ConnectError if the Ollama daemon isn't running
        print(f"Warning: Could not check Ollama model status. Error: {e}")
        for model_name in agent_models_to_check:
            agent_status[model_name] = {"status": "unavailable", "error": "Ollama service not running"}

    # --- Check Whisper Models ---
    try:
        whisper_cache_path = os.path.join(os.path.expanduser("~"), ".cache", "whisper")
        os.makedirs(whisper_cache_path, exist_ok=True)
        for model_name in whisper_models_to_check:
            model_file = os.path.join(whisper_cache_path, f"{model_name}.pt")
            whisper_status[model_name] = {"status": "downloaded" if os.path.exists(model_file) else "not_downloaded"}
    except Exception as e:
        print(f"Warning: Could not check Whisper model status. Error: {e}")
        for model_name in whisper_models_to_check:
            whisper_status[model_name] = {"status": "unavailable", "error": "File system error"}

    return jsonify({
        "agent_models": agent_status,
        "whisper_models": whisper_status
    })

@app.route('/api/models/agent/<path:model_name>', methods=['DELETE'])
def delete_agent_model(model_name):
    """Deletes a locally cached Ollama model."""
    try:
        print(f"INFO: Deleting Ollama model '{model_name}'...")
        ollama.delete(model_name)
        print(f"INFO: Model '{model_name}' deleted successfully.")
        return jsonify({"message": f"Model '{model_name}' deleted successfully."}), 200
    except Exception as e:
        # Handle cases where the model doesn't exist or Ollama service is down
        error_message = f"Failed to delete Ollama model '{model_name}': {str(e)}"
        print(f"ERROR: {error_message}")
        return jsonify({"error": error_message}), 500

@app.route('/api/models/whisper/<model_name>', methods=['DELETE'])
def delete_whisper_model(model_name):
    """Deletes a locally cached Whisper model file."""
    try:
        whisper_cache_path = os.path.join(os.path.expanduser("~"), ".cache", "whisper")
        model_file = os.path.join(whisper_cache_path, f"{model_name}.pt")
        if os.path.exists(model_file):
            print(f"INFO: Deleting Whisper model '{model_name}' from '{model_file}'...")
            os.remove(model_file)
            print(f"INFO: Model '{model_name}' deleted successfully.")
            return jsonify({"message": f"Whisper model '{model_name}' deleted."}), 200
        return jsonify({"error": "Model file not found."}), 404
    except Exception as e:
        return jsonify({"error": f"Failed to delete Whisper model: {str(e)}"}), 500

@app.route('/api/cache-models', methods=['POST'])
def cache_models():
    """Streams the progress of caching the default models."""
    data = request.json
    agent_models = data.get('agent_models', [])
    whisper_models = data.get('whisper_models', [])

    if not agent_models and not whisper_models:
        return Response(json.dumps({'status': 'error', 'message': 'No models selected for caching.'}), mimetype='application/json', status=400)

    def generate_cache_progress():
        try:
            total_models = len(agent_models) + len(whisper_models)
            completed_models = 0

            # 1. Pull Ollama models
            for agent_model in agent_models:
                yield f"data: {json.dumps({'status': 'in_progress', 'message': f'Pulling agent model: {agent_model}...', 'progress': 0})}\n\n"
                try:
                    for progress in ollama.pull(agent_model, stream=True):
                        percentage = 0
                        # Ensure 'total' and 'completed' exist and are not None before calculating percentage
                        total = progress.get("total")
                        completed = progress.get("completed")
                        if total is not None and completed is not None and total > 0:
                            percentage = round((completed / total) * 100)
                        status_message = progress.get('status', f'Pulling {agent_model}...')
                        if 'completed' in progress and 'total' in progress:
                            status_message = f"Downloading: {round(progress['completed']/1e9, 2)}GB / {round(progress['total']/1e9, 2)}GB"
                        yield f"data: {json.dumps({'status': 'in_progress', 'message': status_message, 'progress': percentage})}\n\n"
                    completed_models += 1
                except Exception as e:
                    yield f"data: {json.dumps({'status': 'error', 'message': f'Failed to pull {agent_model}: {str(e)}'})}\n\n"

            # 2. Load Whisper models (which also downloads if not present)
            for whisper_model_name in whisper_models:
                yield f"data: {json.dumps({'status': 'in_progress', 'message': f'Caching Whisper model: {whisper_model_name}...', 'progress': 100})}\n\n"
                get_whisper_model(whisper_model_name)
                completed_models += 1

            yield f"data: {json.dumps({'status': 'complete', 'message': f'Successfully cached {completed_models}/{total_models} selected models.'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'status': 'error', 'message': str(e)})}\n\n"
    return Response(generate_cache_progress(), mimetype='text/event-stream')

@app.route('/api/clear-gpu-cache', methods=['POST'])
def clear_gpu_cache():
    """Endpoint to manually clear the CUDA cache and unload models."""
    if DEVICE == 'cuda':
        print("INFO: Clearing GPU cache and unloading all models.")
        loaded_models.clear()
        torch.cuda.empty_cache()
        # The default model will be reloaded on the next analysis request.
        return jsonify({"message": "GPU cache cleared and all models unloaded."}), 200
    return jsonify({"message": "Not using GPU, no cache to clear."}), 200

@app.route('/api/system-info')
def system_info():
    """Provides information about the system's CPU and GPU."""
    gpu_info = {}
    if DEVICE == 'cuda' and torch.cuda.is_available():
        torch.cuda.synchronize()
        free, total = torch.cuda.mem_get_info()
        used = total - free
        gpu_info = {
            'name': torch.cuda.get_device_name(0),
            'total_memory': total,
            'free_memory': free,
            'percent': round((used / total) * 100, 1) if total > 0 else 0,
        }

    v_mem = psutil.virtual_memory()
    disk = psutil.disk_usage('/')

    return jsonify({
        'device': DEVICE,
        'gpu': gpu_info,
        'cpu': {
            'percent': psutil.cpu_percent(interval=0.1),
        },
        'ram': {
            'total': v_mem.total, 'available': v_mem.available, 'percent': v_mem.percent
        },
        'storage': {
            'total': disk.total, 'used': disk.used, 'percent': disk.percent
        }
    })

@app.route('/api/process', methods=['POST'])
def process_files():
    """
    API endpoint to process an audio file and a screenshot.
    Streams progress updates using Server-Sent Events.
    """
    if 'audio' not in request.files and 'screenshot' not in request.files:
        return jsonify({"error": "Please upload at least one file (audio or screenshot)."}), 400

    # Load defaults from the unified model config
    app_config = load_model_names()
    model_name = request.form.get('model', app_config.get('default_whisper', 'base'))
    agent_model = request.form.get('agent_model', app_config.get('default_agent_model', 'llama3'))
    require_human_review = str(request.form.get('require_human_review', 'false')).strip().lower() in ('1', 'true', 'yes', 'on')
    approval_granted = str(request.form.get('approve_actions', 'false')).strip().lower() in ('1', 'true', 'yes', 'on')

    screenshot_file = request.files.get('screenshot')
    audio_file = request.files.get('audio')

    # Use unique filenames to avoid race conditions
    audio_path = None
    screenshot_path = None

    if audio_file and audio_file.filename:
        audio_filename = f"{uuid.uuid4()}_{audio_file.filename}"
        audio_path = os.path.join(app.config['UPLOAD_FOLDER'], audio_filename)
        audio_file.save(audio_path)

    if screenshot_file and screenshot_file.filename:
        screenshot_filename = f"{uuid.uuid4()}_{screenshot_file.filename}"
        screenshot_path = os.path.join(app.config['UPLOAD_FOLDER'], screenshot_filename)
        screenshot_file.save(screenshot_path)

    def generate_progress():
        try:
            if screenshot_path:
                yield f"data: {json.dumps({'step': 'ocr', 'status': 'in_progress', 'message': 'Extracting text from image with EasyOCR...'})}\n\n"
            else:
                yield f"data: {json.dumps({'step': 'ocr', 'status': 'skipped'})}\n\n"

            if audio_path:
                yield f"data: {json.dumps({'step': 'transcribe', 'status': 'in_progress', 'message': f'Transcribing with Whisper ({model_name})...'})}\n\n"
            else:
                yield f"data: {json.dumps({'step': 'transcribe', 'status': 'skipped'})}\n\n"

            yield f"data: {json.dumps({'step': 'summarize', 'status': 'in_progress', 'message': f'Generating AI summary with {agent_model}...'})}\n\n"
            yield f"data: {json.dumps({'step': 'classify', 'status': 'in_progress', 'message': 'Classifying issue type for routed agent handling...'})}\n\n"
            yield f"data: {json.dumps({'step': 'action', 'status': 'in_progress', 'message': 'Preparing category action plan and approval check...'})}\n\n"

            initial_state: AgentState = {
                **build_initial_state(
                    audio_path=audio_path,
                    screenshot_path=screenshot_path,
                    audio_filename=audio_file.filename if audio_file and audio_file.filename else "N/A",
                    screenshot_filename=screenshot_file.filename if screenshot_file and screenshot_file.filename else "N/A",
                    model_name=model_name,
                    agent_model=agent_model,
                    require_human_review=require_human_review,
                    approval_granted=approval_granted,
                    persist_history=True,
                )
            }

            for update in ANALYSIS_GRAPH.stream(initial_state, stream_mode="updates"):
                if "ocr" in update and screenshot_path:
                    yield f"data: {json.dumps({'step': 'ocr', 'status': 'complete', 'result': update['ocr'].get('ocr_result', [])})}\n\n"
                if "transcribe" in update and audio_path:
                    yield f"data: {json.dumps({'step': 'transcribe', 'status': 'complete', 'result': update['transcribe'].get('transcription_text', '')})}\n\n"
                if "classify_issue" in update:
                    yield f"data: {json.dumps({'step': 'classify', 'status': 'complete', 'result': update['classify_issue'].get('issue_category', 'Other Issues'), 'specialist': update['classify_issue'].get('specialist_path', 'General support specialist')})}\n\n"
                if "summarize_billing" in update:
                    yield f"data: {json.dumps({'step': 'summarize', 'status': 'complete', 'result': update['summarize_billing'].get('summary_text', '')})}\n\n"
                if "summarize_connectivity" in update:
                    yield f"data: {json.dumps({'step': 'summarize', 'status': 'complete', 'result': update['summarize_connectivity'].get('summary_text', '')})}\n\n"
                if "summarize_general" in update:
                    yield f"data: {json.dumps({'step': 'summarize', 'status': 'complete', 'result': update['summarize_general'].get('summary_text', '')})}\n\n"
                if "review_gate" in update:
                    needs_review = update['review_gate'].get('needs_human_review', False)
                    status = 'action_required' if needs_review else 'complete'
                    message = update['review_gate'].get('review_reason', 'No manual review required.')
                    yield f"data: {json.dumps({'step': 'review', 'status': status, 'message': message, 'required': needs_review})}\n\n"
                if "human_review_note" in update:
                    yield f"data: {json.dumps({'step': 'review', 'status': 'complete', 'message': 'Human review note appended to summary output.'})}\n\n"
                if "propose_action" in update:
                    yield f"data: {json.dumps({'step': 'action', 'status': 'complete', 'result': update['propose_action'].get('proposed_action', '')})}\n\n"
                if "approval_gate" in update:
                    approval_status = update['approval_gate'].get('action_status', 'awaiting_approval')
                    msg = update['approval_gate'].get('action_result', '')
                    status = 'action_required' if approval_status == 'awaiting_approval' else 'complete'
                    yield f"data: {json.dumps({'step': 'action', 'status': status, 'result': approval_status, 'message': msg})}\n\n"
                if "execute_action" in update:
                    yield f"data: {json.dumps({'step': 'action', 'status': 'complete', 'result': update['execute_action'].get('action_status', ''), 'message': update['execute_action'].get('action_result', '')})}\n\n"

            yield f"data: {json.dumps({'step': 'done'})}\n\n"
        except Exception as e:
            yield f"data: {json.dumps({'error': str(e)})}\n\n"
        finally:
            # 5. Clean up uploaded files
            if audio_path and os.path.exists(audio_path):
                os.remove(audio_path)
            if screenshot_path and os.path.exists(screenshot_path):
                os.remove(screenshot_path)
            
            # 6. Clear GPU cache and unload models after analysis
            if DEVICE == 'cuda':
                yield f"data: {json.dumps({'step': 'cleanup', 'status': 'in_progress', 'message': 'GPU cache cleared and all models unloaded.'})}\n\n"
                loaded_models.clear()
                torch.cuda.empty_cache()

    return Response(generate_progress(), mimetype='text/event-stream')

if __name__ == '__main__':
    init_customer_database()
    if os.environ.get('WERKZEUG_RUN_MAIN') == 'true' or not app.debug:
        start_customer_monitor()
    app.run(host='0.0.0.0', port=5000, debug=True, threaded=True)